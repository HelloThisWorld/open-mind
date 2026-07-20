"""XLSX (Office Open XML spreadsheet) via ``openpyxl``.

FORMULAS ARE READ AS TEXT, NEVER EVALUATED
------------------------------------------
``openpyxl`` is opened with ``data_only=False``, which yields the formula
*string* (``=SUM(A1:A9)``) rather than a cached value. That is deliberate and it
is the safe direction: evaluating a formula would mean executing untrusted
expressions, and trusting a cached value would mean reporting a number the file
merely claims. The formula text is the fact the document actually contains, so
that is what becomes a Segment — recorded in the cell's metadata as a formula so
no consumer mistakes it for a computed result.

External workbook links are never followed, macros are never read (``.xlsm`` is
not claimed at all), and charts and images are recorded as unsupported content.

SAFETY
------
Same two-stage gate as DOCX: the ZIP package is validated for member count,
size, expansion ratio and path traversal before ``openpyxl`` sees it, and
``defusedxml.defuse_stdlib()`` must be in force or the parser declines.

BOUNDS
------
``read_only=True`` streams rows instead of materializing the workbook, and three
independent caps apply — sheets, rows per sheet, and total cells. Whichever binds
first yields ``partial`` plus a warning naming that exact limit.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional, Tuple

from ..domain.types import ContentMode, DocumentBlockType, DocumentParseStatus
from .builder import DocumentBuilder
from .models import (DocumentParseContext, DocumentProbe, ParsedDocument,
                     dependency_unavailable)
from .probe import _XLSX_MARKER
from .security import DocumentSecurityError, harden_xml, inspect_zip

_MEDIA = ("application/vnd.openxmlformats-officedocument"
          ".spreadsheetml.sheet")

#: A macro-enabled workbook is NOT claimed in this phase. Claiming it would mean
#: opening a file whose whole point is executable content, for no extra value —
#: the cells are the same, and refusing is the honest, bounded answer.
_MACRO_MARKER = "xl/vbaProject.bin"


def CLAIMS(probe: DocumentProbe) -> bool:      # noqa: N802 - registry protocol
    """Claimed on PACKAGE STRUCTURE (``xl/workbook.xml``), never the suffix."""
    if not (probe.is_zip_package and probe.has_zip_member(_XLSX_MARKER)):
        return False
    if probe.has_zip_member(_MACRO_MARKER) or probe.extension == ".xlsm":
        return False
    return True


class XlsxParser:
    """Sheets, bounded used ranges, rows and merged-cell metadata."""

    name = "xlsx"
    version = "1.0"

    def supports(self, probe: DocumentProbe) -> bool:
        return CLAIMS(probe)

    def parse(self, content: bytes,
              context: DocumentParseContext) -> ParsedDocument:
        limits = context.limits
        doc = ParsedDocument(parser_name=self.name, parser_version=self.version,
                             media_type=_MEDIA,
                             title=context.filename or context.logical_key)

        if not harden_xml():
            return dependency_unavailable(context.filename, self.name,
                                          "defusedxml")
        try:
            import openpyxl
        except ImportError:
            return dependency_unavailable(context.filename, self.name,
                                          "openpyxl")

        try:
            package = inspect_zip(
                content, max_members=limits.zip_max_members,
                max_total_bytes=limits.zip_max_total_bytes,
                max_member_bytes=limits.zip_max_member_bytes,
                max_ratio=limits.zip_max_ratio)
        except DocumentSecurityError as exc:
            doc.status = DocumentParseStatus.UNSUPPORTED
            doc.reason = exc.code
            doc.add_warning(exc.code, exc.message, **exc.detail)
            return doc

        for prefix, kind, detail in (
                ("xl/media/", "embedded-image",
                 "embedded images are recorded but not extracted"),
                ("xl/charts/", "chart",
                 "charts are recorded but not extracted")):
            count = sum(1 for m in package["members"] if m.startswith(prefix))
            if count:
                doc.note_unsupported(kind, count, detail)

        # data_only=False -> formulas as text (never evaluated).
        # read_only=True  -> streamed rows, so a huge workbook cannot be
        #                    materialized into memory before the caps apply.
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False,
                                          read_only=True, keep_links=False)
        try:
            _read_properties(workbook, doc)
            builder = DocumentBuilder(doc, max_blocks=limits.max_blocks,
                                      max_block_chars=limits.max_block_chars)
            key = context.logical_key
            builder.add_root(doc.title, {"kind": "spreadsheet-range",
                                         "document": key, "sheet": "",
                                         "range": "A1"})
            merged = _merged_ranges(content, workbook)
            _read_sheets(workbook, builder, key, doc, limits, merged)
        finally:
            try:
                workbook.close()
            except Exception:
                pass
        return doc


def _read_properties(workbook: Any, doc: ParsedDocument) -> None:
    try:
        props = workbook.properties
    except Exception:
        return

    def _text(value: Any) -> str:
        if value is None:
            return ""
        return value.isoformat() if hasattr(value, "isoformat") else str(value)

    doc.metadata.author = _text(getattr(props, "creator", ""))
    doc.metadata.created = _text(getattr(props, "created", None))
    doc.metadata.modified = _text(getattr(props, "modified", None))
    revision = getattr(props, "revision", None)
    if revision:
        doc.metadata.version_label = str(revision)
    title = _text(getattr(props, "title", ""))
    if title.strip():
        doc.title = title.strip()


def _merged_ranges(content: bytes, workbook: Any) -> Dict[str, List[str]]:
    """sheet title -> merged range strings.

    ``read_only`` worksheets do not expose ``merged_cells``, so the workbook is
    reopened non-read-only ONLY for this metadata. It is worth the second pass:
    a merged header cell is exactly the kind of structure that makes a
    spreadsheet readable, and silently losing it would misrepresent the sheet.
    Failure here is non-fatal — merge metadata is an enrichment, not the content.
    """
    try:
        import openpyxl
        book = openpyxl.load_workbook(io.BytesIO(content), data_only=False,
                                      read_only=False, keep_links=False)
    except Exception:
        return {}
    out: Dict[str, List[str]] = {}
    try:
        for sheet in book.worksheets:
            ranges = [str(r) for r in getattr(sheet, "merged_cells", []).ranges] \
                if getattr(sheet, "merged_cells", None) else []
            if ranges:
                out[sheet.title] = sorted(ranges)
    except Exception:
        return out
    finally:
        try:
            book.close()
        except Exception:
            pass
    return out


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters or "A"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _looks_like_header(cells: List[str]) -> bool:
    """A first row is a header when every populated cell is non-numeric text.
    Anything less certain is data, and columns stay positional."""
    populated = [c for c in cells if c.strip()]
    if len(populated) < 2:
        return False
    for cell in populated:
        try:
            float(cell.replace(",", ""))
            return False
        except ValueError:
            continue
    return True


def _read_sheets(workbook: Any, builder: DocumentBuilder, key: str,
                 doc: ParsedDocument, limits: Any,
                 merged: Dict[str, List[str]]) -> None:
    sheet_names = list(workbook.sheetnames)
    hidden = 0
    total_cells = 0
    sheets_read = 0
    cells_capped = False

    for sheet_index, name in enumerate(sheet_names, start=1):
        if builder.full:
            break
        if sheets_read >= limits.xlsx_max_sheets:
            doc.mark_partial(
                "limit-exceeded",
                f"the workbook has {len(sheet_names)} sheets, above "
                f"XLSX_MAX_SHEETS ({limits.xlsx_max_sheets}); the remainder was "
                f"not extracted",
                limit="XLSX_MAX_SHEETS", observed=len(sheet_names),
                allowed=limits.xlsx_max_sheets)
            break
        sheet = workbook[name]
        if getattr(sheet, "sheet_state", "visible") != "visible":
            # A hidden sheet is content the author removed from view. It is
            # counted so its absence is visible, not silently indexed.
            hidden += 1
            continue
        sheets_read += 1
        total_cells, cells_capped = _read_one_sheet(
            sheet, name, sheet_index, builder, key, doc, limits, merged,
            total_cells, cells_capped)

    if hidden:
        doc.note_unsupported("hidden-sheet", hidden,
                             "hidden sheets are not indexed")
    doc.coverage["sheets"] = len(sheet_names)
    doc.coverage["sheets_read"] = sheets_read
    doc.coverage["cells"] = total_cells


def _read_one_sheet(sheet: Any, name: str, sheet_index: int,
                    builder: DocumentBuilder, key: str, doc: ParsedDocument,
                    limits: Any, merged: Dict[str, List[str]],
                    total_cells: int, cells_capped: bool) -> Tuple[int, bool]:
    sheet_key = f"sheet-{sheet_index:03d}"
    sheet_merges = merged.get(name, [])
    builder.add(
        DocumentBlockType.SHEET, name, key=sheet_key,
        locator={"kind": "spreadsheet-range", "document": key, "sheet": name,
                 "range": f"A1:{_column_letter(max(1, sheet.max_column or 1))}"
                          f"{max(1, sheet.max_row or 1)}"},
        indexable=False, content_mode=ContentMode.DERIVED,
        metadata={"sheet": name, "index": sheet_index,
                  "merged_ranges": ", ".join(sheet_merges),
                  "merged_count": len(sheet_merges)})

    headers: List[str] = []
    rows_read = 0
    formulas = 0
    rows_capped = False

    for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if builder.full:
            break
        if rows_read >= limits.xlsx_max_rows_per_sheet:
            if not rows_capped:
                rows_capped = True
                doc.mark_partial(
                    "limit-exceeded",
                    f"sheet {name!r} has more than XLSX_MAX_ROWS_PER_SHEET "
                    f"({limits.xlsx_max_rows_per_sheet}) rows; the remainder "
                    f"was not extracted",
                    limit="XLSX_MAX_ROWS_PER_SHEET", sheet=name,
                    allowed=limits.xlsx_max_rows_per_sheet)
            break
        if total_cells >= limits.xlsx_max_cells:
            if not cells_capped:
                cells_capped = True
                doc.mark_partial(
                    "limit-exceeded",
                    f"the workbook reached XLSX_MAX_CELLS "
                    f"({limits.xlsx_max_cells}); the remaining cells were not "
                    f"extracted",
                    limit="XLSX_MAX_CELLS", allowed=limits.xlsx_max_cells)
            break

        values = list(row or ())
        total_cells += len(values)
        cells = [_cell_text(v) for v in values]
        row_formulas = [i for i, v in enumerate(values) if _is_formula(v)]
        formulas += len(row_formulas)
        if not any(c.strip() for c in cells):
            continue
        rows_read += 1

        # A header row is ALSO emitted as a block, using positional column
        # letters. Emitting only the rows below it loses the whole content of a
        # sheet whose single populated row happens to look like a header — and
        # a header row is real content anyone might search for.
        is_header_row = not headers and _looks_like_header(cells)
        if is_header_row:
            headers = [c.strip() for c in cells]

        names = ([_column_letter(i + 1) for i in range(len(cells))]
                 if is_header_row else headers)
        pairs = [f"{names[i] if i < len(names) and names[i] else _column_letter(i + 1)}"
                 f": {value}"
                 for i, value in enumerate(cells) if value.strip()]
        last_col = _column_letter(max(1, len(cells)))
        builder.add(
            DocumentBlockType.CELL_RANGE, "; ".join(pairs),
            key=f"{sheet_key}-r{row_no:06d}",
            locator={"kind": "spreadsheet-range", "document": key,
                     "sheet": name, "range": f"A{row_no}:{last_col}{row_no}"},
            # DERIVED: `column: value` is a rendering of the row, not the cell
            # text as stored. The locator still cites the exact range.
            content_mode=ContentMode.DERIVED, parent_key=sheet_key,
            metadata={"sheet": name, "row": row_no, "cells": len(cells),
                      "header_row": is_header_row,
                      "formula_columns": ", ".join(
                          _column_letter(i + 1) for i in row_formulas),
                      # Recorded so no consumer mistakes a formula STRING for a
                      # computed result: nothing here is ever evaluated.
                      "has_formula": bool(row_formulas)})

    if formulas:
        # Recorded so a reader knows these are expressions, not results.
        doc.metadata.extra.setdefault("formula_cells", 0)
        doc.metadata.extra["formula_cells"] += formulas
    return total_cells, cells_capped


__all__ = ["XlsxParser", "CLAIMS"]
